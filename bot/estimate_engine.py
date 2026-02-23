import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging
import re
import os

# Импорты из пакета bot# Старый импорт
# from bot.style_config import StyleConfig

# Новый импорт
from bot.filename_generator import generate_estimate_filename

# Старый вызов
# filename = StyleConfig.generate_filename(p_info, is_internal)

# Новый вызов
filename = generate_estimate

from bot.header_builder import HeaderBuilder


class EstimateBuilder:
    def create_excel(self, raw_data, output_dir, p_info, is_internal=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Смета"

        brand_key = p_info.get('company', 'default')
        style = StyleConfig.get_style(brand_key)
        start_row = HeaderBuilder.apply(ws, p_info, style, is_internal)

        headers = ["№", "Наименование", "Ед.", "Кол-во", "Дни/Часы", "Цена"]
        headers += ["Сумма (Выручка)", "Себест. ед.", "Итого Себест.", "Маржа"] if is_internal else ["Сумма"]

        for col, text in enumerate(headers, 2):
            cell = ws.cell(row=start_row, column=col, value=text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        clean_content = re.sub(r'```[a-zA-Z]*\n?', '', raw_data).strip()
        lines = clean_content.split('\n')
        
        current_row = start_row + 1
        item_no = 1
        last_block_start = start_row + 1 
        
        # Списки для хранения адресов ячеек
        section_total_refs = []  # Ячейки "Итого за раздел"
        ref_sum_services = None   # Ячейка "СУММА УСЛУГ"
        ref_overhead = None       # Ячейка "Накладные расходы"
        ref_taxes = None          # Ячейка "Налоги"

        def clean_num(s):
            if not s or str(s).strip() in ['-', '', 'None']: return 0.0
            s = re.sub(r'[^\d.,-]', '', str(s))
            if not s: return 0.0
            if ',' in s and '.' in s: s = s.replace(',', '')
            else: s = s.replace(',', '.')
            try: return float(s)
            except: return 0.0

        for line in lines:
            line = line.strip()
            if not line or '|' not in line or "Наименование" in line or "---" in line:
                continue
            
            parts = [p.strip().replace('*', '') for p in line.split('|')]
            if parts[0] == '': parts.pop(0)
            if parts and parts[-1] == '': parts.pop(-1)
            if not parts or len(parts) < 2: continue
            
            name = parts[0]
            name_upper = name.upper()
            is_section_header = name.startswith('==')
            # Проверяем, является ли строка итоговой
            is_subtotal = any(word in name_upper for word in ["ИТОГО", "СУММА", "ВСЕГО", "НАЛОГ", "ВОЗНАГРАЖДЕНИЕ", "СБОР"])

            try:
                if is_section_header:
                    end_col = 11 if is_internal else 8
                    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=end_col)
                    cell = ws.cell(row=current_row, column=3, value=name.replace('==', '').strip())
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                    current_row += 1
                    continue

                if not is_subtotal:
                    # ОБЫЧНАЯ СТРОКА УСЛУГИ
                    qty = clean_num(parts[2]) if len(parts) > 2 else 1.0
                    days = clean_num(parts[3]) if len(parts) > 3 else 1.0
                    price = clean_num(parts[4]) if len(parts) > 4 else 0.0
                    
                    ws.cell(row=current_row, column=2, value=item_no)
                    item_no += 1
                    ws.cell(row=current_row, column=3, value=name)
                    ws.cell(row=current_row, column=4, value=parts[1] if len(parts) > 1 else "усл.")
                    ws.cell(row=current_row, column=5, value=qty)
                    ws.cell(row=current_row, column=6, value=days)
                    ws.cell(row=current_row, column=7, value=price).number_format = '#,##0'
                    ws.cell(row=current_row, column=8, value=f"=E{current_row}*F{current_row}*G{current_row}").number_format = '#,##0'
                    
                    if is_internal:
                        cost = clean_num(parts[6]) if len(parts) > 6 else 0.0
                        ws.cell(row=current_row, column=9, value=cost).number_format = '#,##0'
                        ws.cell(row=current_row, column=10, value=f"=E{current_row}*F{current_row}*I{current_row}").number_format = '#,##0'
                        ws.cell(row=current_row, column=11, value=f"=H{current_row}-J{current_row}").number_format = '#,##0'
                
                else:
                    # ЛОГИКА ИТОГОВЫХ СТРОК
                    ws.cell(row=current_row, column=3, value=name).font = Font(bold=True)
                    sum_cell = ws.cell(row=current_row, column=8)
                    sum_cell.font = Font(bold=True)
                    sum_cell.number_format = '#,##0'

                    if "ИТОГО ПО РАЗДЕЛУ" in name_upper:
                        # Суммируем только блок между итогами
                        sum_range = f"H{last_block_start}:H{current_row-1}"
                        sum_cell.value = f"=SUM({sum_range})"
                        section_total_refs.append(f"H{current_row}")
                        last_block_start = current_row + 1 # Сбрасываем начало блока

                    elif "СУММА УСЛУГ" in name_upper:
                        # 1) Суммируем только ячейки "Итого по разделу"
                        if section_total_refs:
                            sum_cell.value = "=" + "+".join(section_total_refs)
                        else:
                            # Если разделов не было, суммируем всё сверху
                            sum_cell.value = f"=SUM(H{start_row+1}:H{current_row-1})"
                        ref_sum_services = f"H{current_row}"
                    
                    elif "ВОЗНАГРАЖДЕНИЕ" in name_upper or "НАКЛАДНЫЕ" in name_upper:
                        # 2) 10% от Суммы услуг
                        base = ref_sum_services if ref_sum_services else f"H{current_row-1}"
                        sum_cell.value = f"={base}*0.1"
                        ref_overhead = f"H{current_row}"
                    
                    elif "НАЛОГ" in name_upper:
                        # 3) Налоги = (Сумма услуг + Накладные)/0,9 - (Сумма услуг + Накладные)
                        s_ref = ref_sum_services if ref_sum_services else "0"
                        o_ref = ref_overhead if ref_overhead else "0"
                        sum_cell.value = f"=(({s_ref}+{o_ref})/0.9)-({s_ref}+{o_ref})"
                        ref_taxes = f"H{current_row}"
                    
                    elif "ИТОГО" in name_upper or "ВСЕГО" in name_upper:
                        # 4) Финальная сумма
                        s_ref = ref_sum_services if ref_sum_services else "0"
                        o_ref = ref_overhead if ref_overhead else "0"
                        t_ref = ref_taxes if ref_taxes else "0"
                        sum_cell.value = f"={s_ref}+{o_ref}+{t_ref}"
                    
                    else:
                        # Обычный Итого (если вдруг встретился вне разделов)
                        sum_range = f"H{last_block_start}:H{current_row-1}"
                        sum_cell.value = f"=SUM({sum_range})"

                    if is_internal:
                        # Во внутренней смете для итогов маржа = Выручка - Себестоимость (если есть колонка J)
                        ws.cell(row=current_row, column=11, value=f"=H{current_row}-J{current_row}").number_format = '#,##0'

                current_row += 1
            except Exception as e:
                logging.error(f"Ошибка в строке {current_row}: {e}")

        # Финализация внешнего вида
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['H'].width = 18

        filename = StyleConfig.generate_filename(p_info, is_internal)
        save_path = output_dir if os.path.isdir(output_dir) else os.path.dirname(output_dir)
        final_path = os.path.join(save_path, filename)
        wb.save(final_path)
        return final_path
