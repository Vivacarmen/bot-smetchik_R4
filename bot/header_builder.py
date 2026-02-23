import os
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

from bot import config


class HeaderBuilder:
    @staticmethod
    def apply(ws, p_info, style, is_internal=False):
        """
        Применяет стилизованную шапку к листу Excel с добавлением логотипа.
        Разделяет данные клиента и добавляет данные менеджера.
        Поддерживает расширенный формат даты и времени.
        """
        brand_key = p_info.get('company', 'default')
        brand_name = style.get('display_name', 'Event Agency')
        brand_color = style.get('color', '2C3E50')
        
        # 1. Добавление логотипа (размер 150x150)
        # Логотипы лежат в папке bot/ рядом с header_builder.py
        logo_path = os.path.join(os.path.dirname(__file__), f"logo_{brand_key}.png")
        
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img.width = 150
                img.height = 150
                # Размещаем логотип в области B2
                ws.add_image(img, 'B2')
                # Сдвигаем текстовые данные ниже логотипа
                title_row = 10
                start_info_row = 12
            except Exception:
                # Если файл поврежден или Pillow не установлен
                title_row = 2
                start_info_row = 4
        else:
            # Если логотипа нет, используем компактные отступы
            title_row = 2
            start_info_row = 4

        # 2. Заголовок документа
        title_text = f"КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ: {brand_name}"
        if is_internal:
            title_text = f"ВНУТРЕННИЙ РАСЧЕТ: {brand_name}"
            
        ws.merge_cells(start_row=title_row, start_column=2, end_row=title_row, end_column=8)
        cell_title = ws.cell(row=title_row, column=2, value=title_text)
        cell_title.font = Font(bold=True, size=14, color=brand_color)
        cell_title.alignment = Alignment(horizontal='left')

        # 3. Информация о проекте (Шапка данных)
        client_person = p_info.get('client_name', 'Не указано')
        client_org = p_info.get('company_name', 'Частное лицо / Компания не указана')
        location = p_info.get('location', 'По согласованию')
        date_time_val = p_info.get('event_date', 'Дата не указана')
        guests = p_info.get('guests_qty', '0')
        manager_info = p_info.get('manager', "Ваш менеджер: @vivacarmen")

        rows_data = [
            ("Контактное лицо:", client_person),
            ("Организация:", client_org),
            ("Локация:", location),
            ("Дата и время:", date_time_val),
            ("Кол-во гостей:", f"{guests} чел."),
            ("Менеджер проекта:", manager_info)
        ]

        for i, (label, value) in enumerate(rows_data):
            row_idx = start_info_row + i
            
            label_cell = ws.cell(row=row_idx, column=2, value=label)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='left')
            
            val_cell = ws.cell(row=row_idx, column=3, value=value)
            val_cell.alignment = Alignment(horizontal='left')
            
            if "Менеджер" in label:
                val_cell.font = Font(italic=True, color=brand_color)

        return start_info_row + len(rows_data) + 1
